#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
import shutil
import requests
from openpyxl import load_workbook


# --------- Workbook Location ---------
WORKBOOK_PATH = Path.home() / "Dropbox" / "Excel Shared" / "Precious Metal Investments.xlsx"
SHEET_NAME = "Current Prices"
# -------------------------------------

def fetch_last_close(symbol: str) -> float:
	urls = [
		f"https://stooq.com/q/d/l/?s={symbol}&i=d",
		f"https://stooq.com/q/l/?s={symbol}&f=sd2t2ohlc&h&e=csv",
	]

	rows = []
	for url in urls:
		try:
			r = requests.get(url, timeout=20)
			r.raise_for_status()
			rows = list(csv.DictReader(r.text.splitlines()))
			if rows:
				break
		except Exception:
			continue

	if not rows:
		raise RuntimeError(f"No data returned for {symbol}")

	close = float(rows[-1]["Close"])

	# Silver sometimes returned in cents
	if symbol.lower() == "si.f" and close > 500:
		close = close / 100.0

	return close


def backup_workbook(path: Path) -> Path:
	stamp = datetime.now().strftime("%Y%m%d%H%M%S")
	dst = path.with_name(f"{path.stem}.backup.{stamp}{path.suffix}")
	shutil.copy2(path, dst)
	return dst


def update_prices(workbook: Path, dry_run: bool = False):
	if not workbook.exists():
		raise FileNotFoundError(f"Workbook not found: {workbook}")

	gold_price = fetch_last_close("gc.f")
	silver_price = fetch_last_close("si.f")
	today = datetime.now().date()

	if dry_run:
		print(f"[DRY RUN]")
		print(f"Gold  -> ${gold_price:.2f}  Date: {today}")
		print(f"Silver-> ${silver_price:.2f}  Date: {today}")
		return gold_price, silver_price

	backup = backup_workbook(workbook)
	print(f"Backup created: {backup}")

	wb = load_workbook(workbook)
	if SHEET_NAME not in wb.sheetnames:
		raise RuntimeError(f"Sheet '{SHEET_NAME}' not found")

	ws = wb[SHEET_NAME]

	# Loop rows to find Gold / Silver in column A
	for row in ws.iter_rows(min_row=2, max_col=3):
		metal_cell = row[0]
		price_cell = row[1]
		date_cell = row[2]

		if not metal_cell.value:
			continue

		metal = str(metal_cell.value).strip().lower()

		if metal == "gold":
			price_cell.value = round(gold_price, 2)
			price_cell.number_format = "$#,##0.00"
			date_cell.value = today
			date_cell.number_format = "m/d/yy"

		elif metal == "silver":
			price_cell.value = round(silver_price, 2)
			price_cell.number_format = "$#,##0.00"
			date_cell.value = today
			date_cell.number_format = "m/d/yy"

	wb.save(workbook)

	print("Updated Current Prices sheet successfully.")
	print(f"Gold: ${gold_price:.2f}")
	print(f"Silver: ${silver_price:.2f}")

	return gold_price, silver_price


def main():
	parser = argparse.ArgumentParser(description="Update Current Prices sheet")
	parser.add_argument(
		"--workbook",
		type=Path,
		default=WORKBOOK_PATH,
		help="Path to workbook",
	)
	parser.add_argument("--dry-run", action="store_true")
	args = parser.parse_args()

	update_prices(args.workbook, dry_run=args.dry_run)


if __name__ == "__main__":
	main()